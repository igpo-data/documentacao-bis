from collections import Counter
from functools import lru_cache
from html import escape
from pathlib import Path
import re

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook


UPDATE_PROCEDURES = [
    (
        "atualizar_455_financeiro()",
        "ft_455, ssw_op455, ssw_op455_complementar, dim_frete_455, "
        "dim_frete_compl, dim_cliente, dim_tempo, dim_unidade",
        "Recria a fato financeira da OP455 com dados de frete, mercadoria, "
        "volumes, datas, clientes e unidades.",
        ""
    ),
    (
        "atualizar_dim_002_cotacao(carga)",
        "dim_002, ssw_op002",
        "Atualiza a dimensão de cotações, podendo fazer carga total ou incremental.",
        ""
    ),
    (
        "atualizar_dim_103(carga)",
        "dim_103, ssw_op103_a",
        "Atualiza a dimensão de coletas com situação, horário, canal de solicitação "
        "e dados operacionais.",
        ""
    ),
    (
        "atualizar_dim_200(qtddias)",
        "dim_200, ssw_op200, dim_tempo, dim_horario_descarga",
        "Atualiza manifestos e tempos de descarga da OP200 para um período retroativo.",
        ""
    ),
    (
        "atualizar_dim_200BKP(carga)",
        "dim_200, ssw_op200, dim_horario_descarga",
        "Versão backup/antiga da carga da OP200.",
        ""
    ),
    (
        "atualizar_dim_441(carga)",
        "dim_441, ssw_op441",
        "Atualiza a dimensão de faturas, cobrança, atraso, prorrogação e liquidação.",
        ""
    ),
    (
        "atualizar_dim_915(carga)",
        "dim_915_a_b, ssw_op915_a, ssw_op915_b",
        "Consolida dados da OP915 A e B, ligando CTRC, CT-e e NF-e.",
        ""
    ),
    (
        "atualizar_dim_cliente()",
        "dim_cliente, ssw_op467, ssw_op583, ssw_op455",
        "Atualiza o cadastro de clientes usando cadastro oficial, CNPJ principal e "
        "clientes movimentados na OP455.",
        ""
    ),
    (
        "atualizar_dim_cliente_bkp()",
        "dim_cliente, ssw_op467, ssw_op583, ssw_op455",
        "Versão backup da carga de clientes.",
        ""
    ),
    (
        "atualizar_dim_cliente_teste()",
        "dim_cliente, ssw_op467, ssw_op583, ssw_op455",
        "Versão de teste da carga de clientes, com processamento paginado.",
        ""
    ),
    (
        "atualizar_dim_frete(qtddias)",
        "dim_frete, ssw_op455, ssw_op455_complementar, ssw_op200",
        "Atualiza a dimensão de frete com dados comerciais, operacionais, manifestos "
        "e classificações de operação.",
        ""
    ),
    (
        "atualizar_dim_frete_bkp(carga)",
        "dim_frete, ssw_op455, ssw_op455_complementar, ssw_op200, dim_tempo",
        "Versão backup da carga da dimensão de frete.",
        ""
    ),
    (
        "atualizar_dim_frete_updated(qtddias)",
        "dim_frete, ssw_op455, ssw_op455_complementar, ssw_op200, dim_tempo",
        "Versão atualizada da carga da dimensão de frete usando intervalo retroativo.",
        ""
    ),
    (
        "atualizar_dim_pendencia_pi(carga)",
        "dim_pi_pendencia, sacflow_items_compensation_procedure",
        "Atualiza pendências/indenizações vindas do SacFlow.",
        ""
    ),
    (
        "atualizar_dim_pi_sacflow(carga)",
        "dim_pi_pendencia, sacflow_items_compensation_procedure",
        "Recarrega a dimensão de pendências PI do SacFlow.",
        ""
    ),
    (
        "atualizar_dim_unidade()",
        "dim_unidade, ssw_op455, centros",
        "Atualiza unidades operacionais, regiões, vínculos, UF e cidade.",
        ""
    ),
    (
        "atualizar_dimensoes_ft_002(carga)",
        "dim_002, ft_002, ssw_op002",
        "Orquestra a atualização da dimensão e fato da OP002.",
        ""
    ),
    (
        "atualizar_ft_915(carga)",
        "ft_915, ssw_op915_a, dim_915_a_b, dim_tempo",
        "Atualiza a fato da OP915 com dados financeiros, pesos, frete e comissões.",
        ""
    ),
    (
        "proc_atualizar_chave_cte_cvl_rte_edi_doccob()",
        "rte_base_edi_doccob, ssw_op915_a, ssw_op915_b",
        "Preenche a chave de acesso do CT-e CVL na base RTE/Doccob usando a chave da NF-e.",
        ""
    ),
    (
        "sp_atualizar_chave_cte_cvl()",
        "rte_base_edi_doccob, ssw_op915_a, ssw_op915_b",
        "Faz a mesma atualização da chave CT-e CVL na base RTE/Doccob.",
        ""
    ),
]

DEDUPLICATION_PROCEDURES = [
    (
        "proc_deletar_duplicada_dim_frete()",
        "dim_frete",
        "Remove registros duplicados da dimensão de fretes.",
        ""
    )
]


DETAILS_FILE = Path(__file__).parent / "data" / "Procedure_BD_corrigida_detalhada.xlsx"


def _procedure_key(name):
    """Remove os parâmetros exibidos na interface para casar com a planilha."""
    return re.sub(r"\s*\(.*\)\s*$", "", name)


@lru_cache(maxsize=1)
def _load_procedure_details():
    """Carrega a planilha versionada e agrupa seus campos por procedure."""
    workbook = load_workbook(DETAILS_FILE, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows())]
    details = {}

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        procedure = row.get("Procedure")
        if procedure:
            details.setdefault(str(procedure), []).append(row)

    workbook.close()
    return details


def _split_sources(value):
    if not value:
        return []
    return [item.strip() for item in re.split(r"\s*[;/]\s*", str(value)) if item.strip()]


def _unique(values):
    return list(dict.fromkeys(value for value in values if value))


def _render_flow_diagram(name, detail_rows):
    sources = _unique(
        source
        for row in detail_rows
        for source in _split_sources(row.get("Origem"))
        if source != "Regra da procedure"
    )
    targets = _unique(str(row["Tabela"]) for row in detail_rows if row.get("Tabela"))

    source_cards = "".join(f"<span>{escape(source)}</span>" for source in sources)
    target_cards = "".join(f"<span>{escape(target)}</span>" for target in targets)
    height = max(230, 118 + max(len(sources), len(targets)) * 32)
    html = f"""
    <div class="flow">
      <section><strong>ORIGENS</strong><div class="cards">{source_cards}</div></section>
      <div class="arrow">&#8594;</div>
      <section class="procedure"><strong>PROCEDURE</strong><span>{escape(name)}</span></section>
      <div class="arrow">&#8594;</div>
      <section><strong>DESTINOS</strong><div class="cards">{target_cards}</div></section>
    </div>
    <style>
      body {{ margin: 0; font-family: Arial, sans-serif; color: #243447; }}
      .flow {{ display: grid; grid-template-columns: 1fr 42px 1.25fr 42px 1fr;
               align-items: center; gap: 8px; padding: 14px; border: 1px solid #dce3ea;
               border-radius: 10px; background: #f8fafc; }}
      section {{ text-align: center; }}
      strong {{ display: block; margin-bottom: 9px; color: #52616f; font-size: 11px;
                letter-spacing: .08em; }}
      .cards {{ display: flex; flex-direction: column; gap: 6px; }}
      span {{ display: block; padding: 6px 9px; border-radius: 6px; background: white;
              border: 1px solid #cbd5e1; font-size: 12px; overflow-wrap: anywhere; }}
      .procedure span {{ padding: 14px 10px; color: white; background: #176b87;
                         border-color: #176b87; font-weight: 700; }}
      .arrow {{ text-align: center; color: #176b87; font-size: 27px; font-weight: 700; }}
    </style>
    """
    components.html(html, height=height, scrolling=False)


def _render_explanation(name, description, detail_rows):
    sources = _unique(
        source
        for row in detail_rows
        for source in _split_sources(row.get("Origem"))
        if source != "Regra da procedure"
    )
    targets = _unique(str(row["Tabela"]) for row in detail_rows if row.get("Tabela"))
    actions = _unique(str(row["Agente"]) for row in detail_rows if row.get("Agente"))
    keys = _unique(str(row["Chave"]) for row in detail_rows if row.get("Chave"))
    observations = Counter(
        str(row["Observações"]) for row in detail_rows if row.get("Observações")
    )
    main_observation = observations.most_common(1)[0][0] if observations else ""

    source_text = ", ".join(f"`{source}`" for source in sources)
    target_text = ", ".join(f"`{target}`" for target in targets)
    action_text = ", ".join(actions)
    key_text = "; ".join(keys)

    st.markdown("**O que acontece**")
    st.write(
        f"{description} No detalhamento da carga, a procedure realiza a operação "
        f"“{action_text}” sobre {len(detail_rows)} campos de {target_text}, consumindo "
        f"dados de {source_text}."
    )
    if key_text:
        st.write(
            f"A associação e a atualização dos registros são controladas pelas seguintes "
            f"chaves ou regras de correspondência: {key_text}."
        )
    if main_observation:
        st.write(main_observation)


def _render_detail_table(detail_rows):
    columns = ["Tabela", "Colunas", "Origem", "Chave", "Observações"]
    available_columns = [
        column for column in columns if any(row.get(column) for row in detail_rows)
    ]
    dataframe = pd.DataFrame(detail_rows)[available_columns].fillna("")
    dataframe = dataframe.rename(columns={"Colunas": "Campo de destino"})
    st.markdown("**Mapeamento de campos**")
    st.dataframe(dataframe, use_container_width=True, hide_index=True)


def _render_procedures(procedures, search_term):
    filtered_procedures = [
        procedure
        for procedure in procedures
        if search_term.casefold() in procedure[0].casefold()
    ]

    if not filtered_procedures:
        st.info("Nenhuma procedure encontrada com esse nome.")
        return

    procedure_details = _load_procedure_details()

    for procedure in filtered_procedures:
        name, tables, description, *_ = procedure
        detail_rows = procedure_details.get(_procedure_key(name), [])
        with st.expander(name):
            st.markdown(f"**Tabelas envolvidas:** {tables}")
            st.markdown(f"**Finalidade:** {description}")
            if detail_rows:
                st.markdown("**Fluxo de dados**")
                _render_flow_diagram(name, detail_rows)
                _render_explanation(name, description, detail_rows)
                _render_detail_table(detail_rows)
            else:
                st.caption("Detalhamento de campos ainda não disponível na planilha.")


def render(search_term=""):
    st.title("Banco de dados")
    st.write("Documentação das procedures utilizadas nas rotinas de dados.")

    update_tab, deduplication_tab = st.tabs(
        ["Procedures de atualização", "Procedures de remoção de duplicadas"]
    )

    with update_tab:
        _render_procedures(UPDATE_PROCEDURES, search_term)

    with deduplication_tab:
        _render_procedures(DEDUPLICATION_PROCEDURES, search_term)
